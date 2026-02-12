from dataclasses import dataclass
from pathlib import Path
from typing import Sequence
import pandas as pd
from openpyxl import load_workbook


@dataclass(frozen=True)
class LocalINPNPaths:
    znieff_espece: Path
    znieff_habitats: Path
    znieff_habitats_info: Path
    znieff_infos_generales: Path
    taxref: Path
    n2000_habitats: Path
    n2000_especes_inscrites: Path
    n2000_especes_autres: Path
    n2000_infos_generales: Path
    habref_70: Path

    @staticmethod
    def default(data_dir: Path | str = "data") -> "LocalINPNPaths":
        pdir = Path(data_dir)
        return LocalINPNPaths(
            znieff_espece=pdir / "ZNIEFF_Especes.parquet",
            znieff_habitats=pdir / "ZNIEFF_Habitats.parquet",
            znieff_habitats_info=pdir / "ZNIEFF_Habitats_infos.parquet",
            znieff_infos_generales=pdir / "ZNIEFF_Infos_generales.parquet",
            taxref=pdir / "TAXREFv18.parquet",
            n2000_habitats=pdir / "N2000_Habitats.parquet",
            n2000_especes_inscrites=pdir / "N2000_Especes_inscrites.parquet",
            n2000_especes_autres=pdir / "N2000_Especes_autres.parquet",
            n2000_infos_generales=pdir / "N2000_Infos_generales.parquet",
            habref_70=pdir / "HABREF_70.parquet",
        )


def ensure_exists(path: Path) -> None:
    """Vérifie qu'un fichier existe, lève une exception sinon."""
    if not path.exists():
        raise FileNotFoundError(f"Fichier introuvable : {path}")


def filter_parquet(
    parquet_file: Path,
    key_col: str,
    keep_cols: Sequence[str],
    codes: Sequence[str],
) -> pd.DataFrame:
    """
    Lit uniquement keep_cols depuis le parquet, puis filtre key_col ∈ codes.
    """
    ensure_exists(parquet_file)

    try:
        df = pd.read_parquet(parquet_file, columns=list(keep_cols))
    except ImportError as e:
        raise ImportError(
            "Lecture parquet impossible: aucun moteur parquet disponible. "
            "Installez 'pyarrow' (recommandé) ou 'fastparquet', puis relancez."
        ) from e

    # Robustesse: on filtre en string (parfois parquet peut typer différemment)
    df[key_col] = df[key_col].astype(str)
    code_set = set(str(c) for c in codes)

    df = df[df[key_col].isin(code_set)]
    return df

def write_excel_output(
    out_xlsx: Path,
    df_habitats_znieff: pd.DataFrame | None = None,
    df_especes_znieff: pd.DataFrame | None = None,
    df_habitats_n2000: pd.DataFrame | None = None,
    df_especes_n2000: pd.DataFrame | None = None,
    sheet_habitats_znieff: str = "Habitats ZNIEFF",
    sheet_especes_znieff: str = "Espèces ZNIEFF",
    sheet_habitats_n2000: str = "Habitats N2000",
    sheet_especes_n2000: str = "Espèces N2000",
) -> Path:
    """Écrit les dataframes (même vides) dans un fichier Excel avec plusieurs onglets et ajuste la largeur des colonnes."""

    znieff_habitats_cols = [
        "ID ZNIEFF", "Nom ZNIEFF", "Type ZNIEFF", "Type habitat", "CD_HAB", "Code typologie", "Libellé typologie",
        "Code EUNIS", "Libellé EUNIS", "Code Corine", "Libellé Corine", "Code HIC", "Libellé HIC",
    ]
    znieff_especes_cols = [
        "ID ZNIEFF", "Nom ZNIEFF", "Type ZNIEFF", "Groupe taxonomique", "Nom scientifique", "CD_REF", "CD_NOM", "Type espèce",
    ]
    n2000_habitats_cols = [
        "ID N2000", "Nom site", "Type de zone", "Code HIC", "Libellé HIC", "Forme prioritaire", "CD_HAB",
    ]
    n2000_especes_cols = [
        "ID N2000", "Nom site", "Type de zone", "Groupe taxonomique", "Nom scientifique", "CD_NOM", "CD_REF", "Type espèce",
    ]

    def ensure_headers(df: pd.DataFrame | None, expected_cols: list[str]) -> pd.DataFrame:
        """Garantit la présence des en-têtes attendus.

        - Si `df` est `None`, renvoie un DataFrame vide avec les colonnes attendues.
        - Si `df` existe, ajoute les colonnes manquantes (valeur vide) puis
          réordonne les colonnes attendues en premier.
        """
        if df is None:
            return pd.DataFrame(columns=expected_cols)

        out_df = df.copy()
        for col in expected_cols:
            if col not in out_df.columns:
                out_df[col] = ""

        extra_cols = [c for c in out_df.columns if c not in expected_cols]
        return out_df[expected_cols + extra_cols]

    df_habitats_znieff = ensure_headers(df_habitats_znieff, znieff_habitats_cols)
    df_especes_znieff = ensure_headers(df_especes_znieff, znieff_especes_cols)
    df_habitats_n2000 = ensure_headers(df_habitats_n2000, n2000_habitats_cols)
    df_especes_n2000 = ensure_headers(df_especes_n2000, n2000_especes_cols)
    
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        df_habitats_znieff.to_excel(writer, sheet_name=sheet_habitats_znieff, index=False)
        df_especes_znieff.to_excel(writer, sheet_name=sheet_especes_znieff, index=False)
        df_habitats_n2000.to_excel(writer, sheet_name=sheet_habitats_n2000, index=False)
        df_especes_n2000.to_excel(writer, sheet_name=sheet_especes_n2000, index=False)

    # Charger le workbook avec openpyxl et ajuster les largeurs de colonnes
    wb = load_workbook(out_xlsx)
    
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # Ajouter un peu de padding et appliquer une largeur min/max raisonnable
            adjusted_width = min(50, max(12, max_length + 2))
            ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(out_xlsx)
    return out_xlsx
